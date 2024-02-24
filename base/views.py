from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.response import Response

from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from rest_framework.exceptions import ValidationError

from . serializers import FileUploadSerializer, ImportedDataSerializer
from .models import ImportedData
# Create your views here.


def validate_file_size(file, valid_file_size=10):
    """Check size of file is equal to valid_file_size or not"""
    # convert bytes to megabytes
    file_size_mb = file.size / (1024 * 1024)
    print(file_size_mb, file.size)
    if file_size_mb > valid_file_size:
        raise ValidationError(f"File size exceeds {valid_file_size}MB.")


def validate_list_with_serializer(data_list):
    """
    Validate a serializer with list data
    if valid save to database otherwise return errors.
    It returns list of error dict with row and error 
    """
    errors = []
    serializer = ImportedDataSerializer()
    for row_id, data_dict in enumerate(data_list, start=1):
        try:
            serializer = ImportedDataSerializer(data=data_dict)
            serializer.is_valid(raise_exception=True)
            serializer.save()
        except ValidationError as e:
            errors.append({"row": row_id, "error": e.detail})
    return errors


class UploadFile(APIView):
    """Upload file apiview"""
    serializer_class = FileUploadSerializer

    def get(self, request):
        """Returns a sample file."""
        print(request.user)
        return excel.make_response_from_a_table(
            ImportedData, "xlsx", file_name='Data'
        )

    def post(self, request):
        """Handles the file uploading process"""
        file = request.FILES.get("file", None)

        if not file:
            return Response({'detail': 'Please upload a valid excel file.'}, status=400)

        # read all file
        if file.name.endswith('.xlsx'):
            # validate the file size
            try:
                validate_file_size(file)
            except ValidationError as e:
                return Response({"errors": e.detail}, status=400)

            imported_data = []

            # load file as workbook
            wb = load_workbook(file)

            # make workbook active
            ws = wb.active

            # get all keys form first row of worksheet
            headers = [cell.value for cell in ws[1]]

            # iterate through all rows from 2nd row
            for rows in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}

                for header, value in zip(headers, rows):
                    row_dict[header] = value
                imported_data.append(row_dict)

            # check validity of imported_data
            validation_errors = validate_list_with_serializer(imported_data)

            # return response if any validation errors occur
            if validation_errors:
                return Response({"errors": validation_errors}, status=400)

            return Response({
                'detail': 'Successfully uploaded the file and saved to database'
            }, status=201)

        return Response({
            'error': 'File must have .xlsx extension'
        }, status=400)
